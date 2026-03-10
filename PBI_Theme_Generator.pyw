#--- [0] IMPORT LIBRARIES, DEFINE FUNCTIONS ---

# pip install openpyxl
from openpyxl import load_workbook
from pathlib import Path
import tkinter as tk
from tkinter import messagebox
import warnings

warnings.filterwarnings("ignore", message="Data Validation extension is not supported")

BASE_DIR = Path(__file__).resolve().parent

# Define function for the pop up notification
def show_popup(title, message):
    root = tk.Tk()
    root.withdraw()
    messagebox.showinfo(title, message)
    root.destroy()

# Define function for adjusting sentiment colors based on the primary theme color
def adjust_color(hex_color, factor):
    # Remove the '#' sign
    hex_color = hex_color.lstrip("#")

    # Convert HEX → RGB
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)

    # Adjust brightness
    if factor > 0:  # lighten
        r = int(r + (255 - r) * factor)
        g = int(g + (255 - g) * factor)
        b = int(b + (255 - b) * factor)
    else:  # darken
        r = int(r * (1 + factor))
        g = int(g * (1 + factor))
        b = int(b * (1 + factor))

    # Clamp values
    r = max(0, min(255, r))
    g = max(0, min(255, g))
    b = max(0, min(255, b))

    # Convert back to HEX
    return f"#{r:02X}{g:02X}{b:02X}"

#--- [1] READ THE CHOSEN PARAMETRS FROM THE EXCEL FILE ---

# File path and worksheet
file_path = BASE_DIR / "Theme_Generator.xlsx"
wb = load_workbook(file_path, data_only=True)
ws = wb["Main settings"]

# Read and store the main variables
my_theme = ws["C2"].value
my_font = ws["C3"].value
my_padding = ws["C4"].value
my_theme_name = ws["C13"].value

# Read and store chosen colors
my_colors = []

for row in range(5, 13):  # C5:C12
    cell = ws[f"C{row}"]
    my_color = cell.fill.start_color.rgb
    
    if my_color and my_color != "00000000":
        hex_color = "#" + my_color[-6:]
    else:
        hex_color = "#808080" # Replace with default grey
    
    my_colors.append(hex_color)

# Store colors in separate variables
my_color1, my_color2, my_color3, my_color4, my_color5, my_color6, my_color7, my_color8 = my_colors

# Safeguard for missing values
if my_theme not in ["Light", "Dark"]:
    my_theme = "Light" # Light as default
else:
    my_theme

if my_font not in ["Segoe UI", "Arial"]:
    my_font = "Segoe UI" # Segoe UI as default
else:
    my_font

if my_padding < 0:
    my_padding = 0 # Min cap at 0
elif my_padding > 50:
    my_padding = 50 # Max cap at 50
elif my_padding is None:
    my_padding = 20 # 20 as default
else:
    my_padding

my_padding = str(my_padding) # Convert to string

if my_theme_name is None:
    my_theme_name = "My Theme"

# Calculate the sentiment colors
if my_theme == "Light":
    my_color_max = adjust_color(my_color1, -0.4)  # 40% darker
    my_color_min = adjust_color(my_color1, 0.4)   # 40% lighter
elif my_theme == "Dark":
    my_color_max = adjust_color(my_color1, 0.4)     # 40% lighter
    my_color_min = adjust_color(my_color1, -0.4)    # 40% darker

#--- [2] GENERATE NEW JSON FILE ---

if my_theme == "Light":
    input_file = BASE_DIR / "JSON/Theme Template - Light.json"
elif my_theme == "Dark":
    input_file = BASE_DIR / "JSON/Theme Template - Dark.json"

output_file = BASE_DIR / f"{my_theme_name}.json"

replacements = {
    "_theme_name": my_theme_name,
    "_font": my_font,
    '"_padding"': my_padding,
    "_Color1": my_color1,
    "_Color2": my_color2,
    "_Color3": my_color3,
    "_Color4": my_color4,
    "_Color5": my_color5,
    "_Color6": my_color6,
    "_Color7": my_color7,
    "_Color8": my_color8,
    "_maximum": my_color_max,
    "_minimum": my_color_min
}

with open(input_file, "r", encoding="utf-8") as f:
    content = f.read()

for key, value in replacements.items():
    content = content.replace(key, value)

with open(output_file, "w", encoding="utf-8") as f:
    f.write(content)

print(f"New Power BI theme named '{my_theme_name}' created!\n\nLocation:\n{output_file}")

show_popup(
    "Power BI Theme Generator",
    f"New Power BI theme named '{my_theme_name}' created!\n\nLocation:\n{output_file}"
)